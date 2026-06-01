"""
Generate an Aimsun control plan text file from boom gate open/close event data.

Input workbook format:
  - One sheet per site (e.g. GE, MO, AS …)
  - Row 68: header row (col A = "Date", col D = "State", col G = "Time", col H = "Operating")
  - State=0: boom closes at Time; State=1: boom opens, Operating = closure duration

Usage examples
--------------
List available sites:
    uv run python src/create_cp_boomtime.py --list-sites

Single date, AM peak:
    uv run python src/create_cp_boomtime.py \\
        --sites GE MO AS \\
        --start 06:00 --end 09:30 \\
        --date 2025/03/10 \\
        --cp-name AM_2025_v1

Average across all weekdays in the workbook:
    uv run python src/create_cp_boomtime.py \\
        --sites GE MO AS SJ WO \\
        --start 06:00 --end 09:30 \\
        --cp-name AM_2025_avg
"""

import argparse
import datetime
import sys
from pathlib import Path

import openpyxl

# ---------------------------------------------------------------------------
# Default paths
# ---------------------------------------------------------------------------
DEFAULT_INPUT = (
    r"C:\Users\IvanVelilla\Mobility Lab Limited"
    r"\Projects - 1034 - Western LX Aimsun"
    r"\New North Road\Data"
    r"\Full Network Summary and Open-Close Data_2025 Analysis.xlsm"
)
DEFAULT_OUTPUT = (
    r"C:\Users\IvanVelilla\Mobility Lab Limited"
    r"\Projects - 1034 - Western LX Aimsun"
    r"\New North Road\Data"
    r"\cp_boomtime_am.txt"
)

# ---------------------------------------------------------------------------
# Site configuration — fill in node/signal IDs for new sites before use
# ---------------------------------------------------------------------------
SITE_CONFIG: dict[str, dict] = {
    "WO": {"name": "Woodward Road",     "node": 10264574, "sig_on": 10264578, "sig_off": 10264579},
    "SJ": {"name": "St Jude Street",   "node": 10266940, "sig_on": 10266945, "sig_off": 10266946},
    "MO": {"name": "Morningside Drive", "node": 10267816, "sig_on": 10267822, "sig_off": 10267823},
    "GE": {"name": "George Street",     "node": 10260974, "sig_on": 10260986, "sig_off": 10260987},
    "AS": {"name": "Asquith Ave",       "node": 89356028, "sig_on": 89356031, "sig_off": 89356032},
    "RO": {"name": "Rossgrove",         "node": 89356004, "sig_on": 89356009, "sig_off": 89356010},
    "CH": {"name": "Chalmers St",       "node": 89375286, "sig_on": 89375292, "sig_off": 89375293},
    "SG": {"name": "Saint Georges St",  "node": 89362210, "sig_on": 89371677, "sig_off": 89375283},
    # New sites — update node/sig_on/sig_off before use
    "FR": {"name": "FR",  "node": 0, "sig_on": 0, "sig_off": 0},
    "GL": {"name": "GL",  "node": 0, "sig_on": 0, "sig_off": 0},
    "SH": {"name": "SH",  "node": 0, "sig_on": 0, "sig_off": 0},
    "BM": {"name": "BM",  "node": 0, "sig_on": 0, "sig_off": 0},
    "ML": {"name": "ML",  "node": 0, "sig_on": 0, "sig_off": 0},
    "MT": {"name": "MT",  "node": 0, "sig_on": 0, "sig_off": 0},
    "SP": {"name": "SP",  "node": 0, "sig_on": 0, "sig_off": 0},
    "MA": {"name": "MA",  "node": 0, "sig_on": 0, "sig_off": 0},
    "TA": {"name": "TA",  "node": 0, "sig_on": 0, "sig_off": 0},
    "WA": {"name": "WA",  "node": 0, "sig_on": 0, "sig_off": 0},
}

# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def _to_sec(t) -> int:
    """Convert datetime.time / datetime.datetime to seconds since midnight."""
    if isinstance(t, (datetime.time, datetime.datetime)):
        return t.hour * 3600 + t.minute * 60 + t.second
    return int(t)


def _parse_time_arg(s: str) -> int:
    """Parse 'HH:MM' or 'HH:MM:SS' argument to seconds since midnight."""
    parts = s.split(":")
    h, m, sec = int(parts[0]), int(parts[1]), int(parts[2]) if len(parts) > 2 else 0
    return h * 3600 + m * 60 + sec


def _fmt_sec(s: int) -> str:
    h, r = divmod(abs(s), 3600)
    m, sc = divmod(r, 60)
    return f"{h:02d}:{m:02d}:{sc:02d}"


# ---------------------------------------------------------------------------
# Reading closures from a sheet
# ---------------------------------------------------------------------------

def _find_header_row(ws) -> int | None:
    """Return 1-based row index of the event header (col A == 'Date')."""
    for i, row in enumerate(ws.iter_rows(min_row=1, max_row=200, values_only=True), start=1):
        if row[0] == "Date":
            return i
    return None


def read_closures(ws, date_filter: set[str] | None = None) -> list[tuple[str, int, int]]:
    """
    Read individual closure events from a sheet.

    Returns list of (date_str, start_sec, duration_sec).
    date_filter: set of 'YYYY/MM/DD' strings; None = all dates.
    """
    header_row = _find_header_row(ws)
    if header_row is None:
        return []

    closures: list[tuple[str, int, int]] = []
    pending_close: tuple[str, int] | None = None  # (date, start_sec) of last State=0

    for row in ws.iter_rows(min_row=header_row + 1, values_only=True):
        if row[0] is None:
            break

        date_str = str(row[0]).strip()[:10]   # "YYYY/MM/DD"
        state = row[3]
        time_val = row[6]    # col G: Time
        operating = row[7]   # col H: Operating (duration, only on State=1)

        if not isinstance(state, (int, float)):
            continue

        if date_filter is not None and date_str not in date_filter:
            pending_close = None
            continue

        if int(state) == 0 and isinstance(time_val, (datetime.time, datetime.datetime)):
            pending_close = (date_str, _to_sec(time_val))

        elif int(state) == 1 and isinstance(operating, (datetime.time, datetime.datetime)):
            dur = _to_sec(operating)
            if pending_close is not None:
                closures.append((pending_close[0], pending_close[1], dur))
            pending_close = None

    return closures


def get_weekday_dates(ws) -> list[str]:
    """Return sorted list of weekday date strings from the sheet's event data."""
    header_row = _find_header_row(ws)
    if header_row is None:
        return []

    seen: set[str] = set()
    result: list[str] = []
    for row in ws.iter_rows(min_row=header_row + 1, values_only=True):
        if row[0] is None:
            break
        date_str = str(row[0]).strip()[:10]
        if date_str in seen:
            continue
        seen.add(date_str)
        try:
            d = datetime.datetime.strptime(date_str, "%Y/%m/%d")
            if d.weekday() < 5:
                result.append(date_str)
        except ValueError:
            pass
    return sorted(set(result))


# ---------------------------------------------------------------------------
# Averaging closures across multiple days
# ---------------------------------------------------------------------------

def average_closures(
    closures: list[tuple[str, int, int]],
    window_start: int,
    window_end: int,
    gap_threshold: int = 900,
) -> list[tuple[int, int]]:
    """
    Group closures that occur at similar times across days (same train slot)
    and return averaged (start_sec, duration_sec) for each group.

    gap_threshold: seconds between consecutive closures to split into a new group.
    """
    filtered = sorted(
        [(s, d) for (_, s, d) in closures if window_start <= s < window_end],
        key=lambda x: x[0],
    )
    if not filtered:
        return []

    groups: list[list[tuple[int, int]]] = []
    current: list[tuple[int, int]] = [filtered[0]]
    for item in filtered[1:]:
        if item[0] - current[-1][0] <= gap_threshold:
            current.append(item)
        else:
            groups.append(current)
            current = [item]
    groups.append(current)

    return [
        (round(sum(s for s, _ in g) / len(g)), round(sum(d for _, d in g) / len(g)))
        for g in groups
    ]


# ---------------------------------------------------------------------------
# Building phase sequence
# ---------------------------------------------------------------------------

def build_phases(
    closures: list[tuple[str, int, int]],
    window_start: int,
    window_end: int,
) -> list[tuple[int, bool]]:
    """
    Convert a list of (date, start_sec, duration_sec) closures into a sequence
    of (duration_sec, is_red) phases that exactly sum to (window_end - window_start).

    is_red=True → boom closed; is_red=False → boom open (green).
    """
    cycle = window_end - window_start

    events = sorted(
        [(s, d) for (_, s, d) in closures if window_start <= s < window_end],
        key=lambda x: x[0],
    )

    phases: list[tuple[int, bool]] = []
    clock = window_start
    total = 0

    for boom_start, dur in events:
        green = round(boom_start - clock)
        red = round(dur)
        if green < 0:
            # This closure overlaps the previous one — skip it
            print(f"  warning: skipping overlapping closure at {_fmt_sec(boom_start)}", file=sys.stderr)
            continue
        if green > 0:
            phases.append((green, False))
            total += green
        phases.append((red, True))
        total += red
        clock = boom_start + dur

    remaining = cycle - total
    if remaining > 0:
        phases.append((remaining, False))
    elif remaining < 0 and phases:
        # Last closure spills past cycle end — trim it
        last_dur, last_is_red = phases[-1]
        trimmed = last_dur + remaining  # remaining is negative
        if trimmed > 0:
            phases[-1] = (trimmed, last_is_red)
        else:
            phases.pop()

    return phases


# ---------------------------------------------------------------------------
# Writing the Aimsun control plan
# ---------------------------------------------------------------------------

def write_cp(
    out_path: Path,
    cp_name: str,
    sites: list[str],
    window_start: int,
    window_end: int,
    closures_per_site: dict[str, list[tuple[str, int, int]]],
    use_average: bool,
    gap_threshold: int = 900,
) -> None:
    cycle = window_end - window_start

    with out_path.open("w", newline="\r\n") as f:
        f.write(cp_name + "\n")

        for site in sites:
            cfg = SITE_CONFIG[site]
            raw = closures_per_site[site]

            if use_average:
                avg = average_closures(raw, window_start, window_end, gap_threshold=gap_threshold)
                # Re-wrap as (date, start, dur) for build_phases
                effective = [("avg", s, d) for s, d in avg]
            else:
                effective = raw

            phases = build_phases(effective, window_start, window_end)
            n = len(phases)
            total_check = sum(d for d, _ in phases)

            print(f"  {site}: {sum(1 for _, r in phases if r)} closures, "
                  f"total={_fmt_sec(total_check)} vs cycle={_fmt_sec(cycle)}", file=sys.stderr)

            f.write(f"{cfg['node']}\n")
            f.write("2\n")                          # fixed signal
            f.write(f"{cycle},0\n")                 # cycle time, offset
            f.write(f"{n},0,4,50,0\n")              # nPhases, ?, yellow, %red, ?

            for dur, is_red in phases:
                sig_id = cfg["sig_off"] if is_red else cfg["sig_on"]
                flag = "True" if is_red else "False"
                f.write(f"{dur},{flag},-1,0,-1,{sig_id}\n")


# ---------------------------------------------------------------------------
# CLI
# ---------------------------------------------------------------------------

def main() -> None:
    parser = argparse.ArgumentParser(
        description="Generate Aimsun boom gate control plan from open/close event data.",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog=__doc__,
    )
    parser.add_argument("--input",    default=DEFAULT_INPUT,  help="Path to source .xlsm workbook")
    parser.add_argument("--output",   default=DEFAULT_OUTPUT, help="Path for output .txt control plan")
    parser.add_argument("--sites",    nargs="+", metavar="SITE",
                        help="Site abbreviations to include (e.g. GE MO AS)")
    parser.add_argument("--start",    default="06:00", help="Window start time HH:MM (default 06:00)")
    parser.add_argument("--end",      default="09:30", help="Window end time   HH:MM (default 09:30)")
    parser.add_argument("--date",     nargs="+", metavar="YYYY/MM/DD",
                        help="One or more specific dates to use. Omit to average all weekdays.")
    parser.add_argument("--cp-name",  default="BoomtimeCP", help="Control plan name (first line of output)")
    parser.add_argument("--gap",      type=int, default=900, metavar="SECONDS",
                        help="Max gap between closures to group as the same train slot when averaging (default 900s)")
    parser.add_argument("--list-sites", action="store_true",
                        help="List available sites in the workbook and exit")
    args = parser.parse_args()

    wb_path = Path(args.input)
    if not wb_path.exists():
        sys.exit(f"Input file not found: {wb_path}")

    print(f"Loading workbook: {wb_path.name} …", file=sys.stderr)
    wb = openpyxl.load_workbook(str(wb_path), read_only=True, keep_vba=False, data_only=True)
    available = [s for s in wb.sheetnames if s in SITE_CONFIG]

    if args.list_sites:
        print("\nAvailable sites in workbook:")
        for s in available:
            cfg = SITE_CONFIG[s]
            status = "ok" if cfg["node"] != 0 else "WARNING: node ID not set"
            print(f"  {s:4s}  {cfg['name']:25s}  node={cfg['node']:12d}  {status}")
        wb.close()
        return

    # Validate requested sites
    if not args.sites:
        sys.exit("Specify sites with --sites (or use --list-sites to see what's available).")

    bad = [s for s in args.sites if s not in SITE_CONFIG]
    if bad:
        sys.exit(f"Unknown site(s): {bad}. Run --list-sites to see valid codes.")

    missing = [s for s in args.sites if s not in wb.sheetnames]
    if missing:
        sys.exit(f"Site(s) not found as sheets in workbook: {missing}")

    unconfigured = [s for s in args.sites if SITE_CONFIG[s]["node"] == 0]
    if unconfigured:
        print(f"WARNING: Aimsun node/signal IDs not set for: {unconfigured}. "
              "Update SITE_CONFIG in this script before use.", file=sys.stderr)

    window_start = _parse_time_arg(args.start)
    window_end   = _parse_time_arg(args.end)
    if window_end <= window_start:
        sys.exit("--end must be later than --start")

    date_filter: set[str] | None = set(args.date) if args.date else None
    use_average = date_filter is None or len(date_filter) > 1

    # Read closures for each site
    closures_per_site: dict[str, list] = {}
    for site in args.sites:
        ws = wb[site]

        if date_filter is None:
            dates = set(get_weekday_dates(ws))
            print(f"  {site}: averaging {len(dates)} weekdays", file=sys.stderr)
        else:
            dates = date_filter
            print(f"  {site}: using date(s) {sorted(dates)}", file=sys.stderr)

        closures_per_site[site] = read_closures(ws, dates if date_filter is None else date_filter)

    wb.close()

    # Write output
    out_path = Path(args.output)
    out_path.parent.mkdir(parents=True, exist_ok=True)
    print(f"\nWriting control plan → {out_path.name}", file=sys.stderr)

    write_cp(
        out_path=out_path,
        cp_name=args.cp_name,
        sites=args.sites,
        window_start=window_start,
        window_end=window_end,
        closures_per_site=closures_per_site,
        use_average=use_average,
    )

    print("Done.", file=sys.stderr)


if __name__ == "__main__":
    main()
